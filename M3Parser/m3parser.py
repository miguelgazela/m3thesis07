import os
from os import walk
import csv
import openpyxl
from m3emailuser import M3EmailUser

class M3Parser(object):

    def __init__(self, path):
        self.path = path

    def executeHousePartyProtocol(self):
        print "   Executing House Party Protocol..."

        # get all user folders
        user_folders = []
        for (dirpath, dirnames, filenames) in walk(self.path):
            user_folders.extend(dirnames)
            break

        print "   Found " + str(len(user_folders)) + " users"

        # get mailboxes for each user
        self.users = []
        for user_folder in user_folders[:1]:

            mailboxes = []

            for (dirpath, dirnames, filenames) in walk(self.path + "/" + user_folder):
                print dirnames
                mailboxes.extend(dirnames)
                break

            self.users.append(M3EmailUser(user_folder, mailboxes))

        for user in self.users:
            print "   Analysing " + user.folder
            user.analyse_emails()
            user.create_analysis_files()

        self.create_statistics_file()
        self.create_exchange_matrix()


    def create_statistics_file(self):

        # create csv file with general statistics

        if not os.path.exists("analysis/"):
            os.makedirs("analysis")

        print "   Saving csv file with statistics "

        f = open("analysis/statistics.csv", "wt")
        try:
            writer = csv.writer(f)
            writer.writerow(('counter', 'username', '#mailboxes', '#emails'))

            counter = 1
            for user in self.users:
                writer.writerow((counter, user.folder, len(user.mailbox_paths), len(user.emails)))
                counter += 1

        finally:
            f.close()


    def create_exchange_matrix(self):

        # exchange_matrix = {
        #     'miguel@gmail.com': {
        #         'barbara@gmail.com': 4,
        #         'andre@gmail.com': 3,
        #         'ademar@gmail.com': 5
        #     },
        #     'barbara@gmail.com': {
        #         'miguel@gmail.com': 9,
        #         'andre@gmail.com': 1,
        #         'ademar@gmail.com': 12,
        #         'joao@gmail.com': 8
        #     },
        #     'andre@gmail.com': {
        #         'barbara@gmail.com': 8,
        #         'tita@gmail.com': 34,
        #         'ademar@gmail.com': 3
        #     },
        #     'ademar@gmail.com': {
        #         'barbara@gmail.com': 1,
        #         'andre@gmail.com': 1,
        #         'miguel@gmail.com': 1,
        #         'ribas@gmail.com': 9
        #     },
        # }

        for user in self.users:

            exchange_matrix = {}

            f = open("analysis/" + user.folder + "/analysis_1.csv", 'rt')
            try:
                reader = csv.reader(f)
                for row in reader:

                    if row[0] != 'from':

                        sender = row[0]
                        if sender is not None and len(sender) > 0:

                            sender_graph = exchange_matrix.get(sender, {})

                            tos = [address.strip() for address in row[1].split(',') if address != '']
                            if len(tos) > 0:

                                for address in tos:
                                    counter = sender_graph.get(address, 0)
                                    counter += 1
                                    sender_graph[address] = counter

                                exchange_matrix[sender] = sender_graph
            finally:
                f.close()

            print "   Saving exchange matrix file"

            f = open("analysis/" + user.folder + "/exchange_matrix.csv", "wt")
            try:
                writer = csv.writer(f)
                writer.writerow(('counter', 'sender', 'receiver', '#msgs'))

                counter = 1
                for sender_key in exchange_matrix.keys():

                    sender_graph = exchange_matrix.get(sender_key)

                    for receiver_key in sender_graph.keys():

                        writer.writerow((counter, sender_key, receiver_key, sender_graph[receiver_key]))
                        counter += 1

            finally:
                f.close()

            # wb = openpyxl.Workbook()
            # sheet = wb.active
            # sheet.title = "Exchange Matrix"
            #
            # receiver_cells = {}
            # last_column = 1
            #
            # counter = 2
            # for sender_key in exchange_matrix.keys():
            #
            #     cell = sheet.cell(row=counter, column=1)
            #     cell.value = unicode(sender_key, errors='ignore')
            #
            #     sender_graph = exchange_matrix.get(sender_key)
            #
            #     for receiver_key in sender_graph.keys():
            #
            #         column = receiver_cells.get(receiver_key, None)
            #
            #         if column is not None:
            #
            #             body_cell = sheet.cell(row=counter, column=column)
            #             body_cell.value = sender_graph[receiver_key]
            #
            #         else:
            #             header_cell = sheet.cell(row=1, column=last_column + 1)
            #             header_cell.value = unicode(receiver_key, errors='ignore')
            #
            #             receiver_cells[receiver_key] = last_column + 1
            #
            #             body_cell = sheet.cell(row=counter, column=last_column + 1)
            #             body_cell.value = sender_graph[receiver_key]
            #
            #             last_column += 1
            #
            #     counter += 1
            #
            # wb.save("analysis/" + user.folder + "/exchange_matrix.xlsx")
