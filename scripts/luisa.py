import openpyxl
import os
from os import walk
from datetime import datetime

divida_wb = openpyxl.load_workbook('divida.xlsx')
lista_wb = openpyxl.load_workbook('lista.xlsx')

sheetD = divida_wb.get_sheet_by_name('Folha1')
sheetL = lista_wb.get_sheet_by_name('Folha1')

row_counter_d = 10

while True:

    client_cell_d = sheetD.cell(row=row_counter_d, column=2)
    before_cell = sheetD.cell(row=row_counter_d, column=17)
    after_cell = sheetD.cell(row=row_counter_d, column=18)

    client_id = client_cell_d.value
    row_counter_l = 2
    has_found = False
    total_value_before = 0
    total_value_after = 0
    limit_date = datetime.strptime('20160101', '%Y%m%d')

    print str(client_id)

    while True:

        client_cell_l = sheetL.cell(row=row_counter_l, column=3)
        client_id_l = client_cell_l.value

        if client_id_l is None:
            break

        # print str(client_id) + " - " + str(client_id_l)

        if client_id == client_id_l:

            # print "SAME"

            has_found = True
            date_cell = sheetL.cell(row=row_counter_l, column=7)
            divida_value = sheetL.cell(row=row_counter_l, column=13)

            date_object = datetime.strptime(str(date_cell.value), "%Y%m%d")

            if date_object < limit_date:

                # if client_id == 1000098:
                #     print str(divida_value.value) + " BEFORE"

                total_value_before += divida_value.value
            else:

                # if client_id == 1000098:
                #     print str(divida_value.value) + " AFTER"

                total_value_after += divida_value.value

        else:

            if has_found:
                break

        row_counter_l += 1


    # if client_id == 1000098:
    #     print "Value: " + str(total_value_before)
    #     print "Value: " + str(total_value_after)

    before_cell.value = total_value_before
    after_cell.value = total_value_after

    row_counter_d += 1

    if client_id is None:
        break

divida_wb.save('divida-done.xlsx')



# print lista.title



            # wb = openpyxl.Workbook()
            # sheet = wb.active
            # sheet.title = "Exchange Matrix"
            #
            # receiver_cells = {}
            # last_column = 1
            #
            # counter = 2
            # for sender_key in exchange_matrix.keys():
            #
            #     cell = sheet.cell(row=counter, column=1)
            #     cell.value = unicode(sender_key, errors='ignore')
            #
            #     sender_graph = exchange_matrix.get(sender_key)
            #
            #     for receiver_key in sender_graph.keys():
            #
            #         column = receiver_cells.get(receiver_key, None)
            #
            #         if column is not None:
            #
            #             body_cell = sheet.cell(row=counter, column=column)
            #             body_cell.value = sender_graph[receiver_key]
            #
            #         else:
            #             header_cell = sheet.cell(row=1, column=last_column + 1)
            #             header_cell.value = unicode(receiver_key, errors='ignore')
            #
            #             receiver_cells[receiver_key] = last_column + 1
            #
            #             body_cell = sheet.cell(row=counter, column=last_column + 1)
            #             body_cell.value = sender_graph[receiver_key]
            #
            #             last_column += 1
            #
            #     counter += 1
            #
            # wb.save("analysis/" + user.folder + "/exchange_matrix.xlsx")
